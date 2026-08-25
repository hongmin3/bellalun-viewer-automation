# -*- coding: utf-8 -*-
"""체크리스트 xlsx에 TC별 자동화 판정 결과를 기록한다.

원본 `Bellalun_Viewer_기본기능_Checklist_개정본.xlsx` 의 TC 행 순서를 그대로 두고
오른쪽에 결과 열만 덧붙여 별도 파일로 저장한다. 원본은 수정하지 않는다.
"""

import os
import shutil
from datetime import datetime

# 기본기능 체크리스트 **개정본**. 이 저장소의 판정 기준 문서다.
#
# 2026-08-19 주의: `지식\(TC) R-23-2346_BellalunViewer_기본기능_Checklist.xlsx`는
# **다른 문서**이고 TC 번호 매핑이 다르다. 그 문서를 기준으로 착각해
# `TC_Basic_WorkFlow_02`를 "범위 불일치"로 잘못 강등한 적이 있다(개정본 WF02는
# "공통 2D/3D 검사 촬영 및 Tool 적용"으로 구현과 정확히 일치한다).
# **판정 결과를 기록하는 원본은 개정본이고, 판정 근거도 개정본에서 확인한다.**
CHECKLIST_NAME = "Bellalun_Viewer_기본기능_Checklist_개정본.xlsx"
CHECKLIST_SHEET = "개정 TC"

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.result import PASS, FAIL, MANUAL, SKIP, BLOCKED, STATUSES

RESULT_HEADERS = ["자동화 판정", "판정 일시", "PASS", "FAIL", "MANUAL", "SKIP", "BLOCKED",
                  "실패 항목", "수동 확인 / 미수행 항목", "증거"]

FILLS = {
    PASS:   PatternFill("solid", fgColor="C6EFCE"),
    FAIL:   PatternFill("solid", fgColor="FFC7CE"),
    MANUAL: PatternFill("solid", fgColor="FFEB9C"),
    SKIP:   PatternFill("solid", fgColor="E7E6E6"),
    BLOCKED: PatternFill("solid", fgColor="E4DFEC"),
    "미수행": PatternFill("solid", fgColor="F2F2F2"),
}
FONTS = {
    PASS:   Font(color="006100", bold=True),
    FAIL:   Font(color="9C0006", bold=True),
    MANUAL: Font(color="9C6500", bold=True),
    SKIP:   Font(color="808080"),
    BLOCKED: Font(color="7030A0", bold=True),
    "미수행": Font(color="A6A6A6"),
}


def source_path(ctx):
    """체크리스트 원본을 PC 독립적으로 찾는다.

    `config.json > checklist_xlsx`가 있으면 그것을 최우선으로 쓰되, **실제로
    존재할 때만** 쓴다. 그렇지 않으면 저장소 상위로 올라가며 `지식` 폴더의
    체크리스트를 찾는다(`core/dbreset.source_dir`와 같은 방식).

    이 함수가 생긴 이유: `config.json`에 다른 PC 사용자의 Downloads 경로가
    박혀 있어 이 PC에서는 파일이 없었고, 결과 기록이 조용히 빠졌다
    (2026-08-18 확인 — 08-10 이후 생성 이력 없음).
    """
    override = (ctx.cfg.get("checklist_xlsx") or "").strip()
    if override and os.path.isfile(override):
        return override
    here = os.path.abspath(ctx.root)
    for _ in range(4):                      # auto -> Bellalun Viewer -> 자동화 ...
        here = os.path.dirname(here)
        if not here:
            break
        for candidate in (os.path.join(here, CHECKLIST_NAME),
                          os.path.join(here, "지식", CHECKLIST_NAME)):
            if os.path.isfile(candidate):
                return candidate
    return override or ""


#: 리포트에 실을 체크리스트 원문 열. 헤더 문구 그대로 찾는다(열 위치를 박지 않는다).
TC_TEXT_COLUMNS = {
    "function": "Function",
    "title": "Title",
    "precondition": "Precondition",
    "steps": "Step Description",
    "expected": "Expected Result",
    "test_data": "Test Data",
}


def read_tc_rows(source_xlsx, sheet_name=None):
    """체크리스트에서 TC 원문을 읽어 `{TC ID: {precondition, steps, ...}}` 로 준다.

    HTML 리포트가 "이 TC 가 무엇을 검증하는가"를 **기준 문서 원문으로** 보여주기
    위한 것이다. 자동화가 요약한 문장이 아니라 원문을 실어야 판정 근거를 감사할
    수 있다(AGENTS.md 0절).

    열은 **헤더 문구로 찾는다** — 열 순서를 박으면 체크리스트가 개정될 때 조용히
    엉뚱한 열을 읽는다.
    """
    if not source_xlsx or not os.path.isfile(source_xlsx):
        return {}
    wb = load_workbook(source_xlsx, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else (
            wb[CHECKLIST_SHEET] if CHECKLIST_SHEET in wb.sheetnames
            else _pick_tc_sheet(wb))
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return {}

    hdr_index = None
    for i, row in enumerate(rows[:20]):
        if any(str(v or "").strip() == "TC ID" for v in row):
            hdr_index = i
            break
    if hdr_index is None:
        return {}
    header = [str(v or "").strip() for v in rows[hdr_index]]
    try:
        tc_col = header.index("TC ID")
    except ValueError:
        return {}
    col_of = {key: (header.index(name) if name in header else None)
              for key, name in TC_TEXT_COLUMNS.items()}

    out = {}
    for row in rows[hdr_index + 1:]:
        tc_id = str(row[tc_col] or "").strip() if tc_col < len(row) else ""
        if not tc_id:
            continue
        entry = {}
        for key, col in col_of.items():
            entry[key] = ("" if col is None or col >= len(row)
                          else str(row[col] or "").strip())
        out[tc_id] = entry
    return out


def _find_header_row(ws, tc_col_name="TC ID"):
    for row in range(1, min(ws.max_row, 20) + 1):
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row, col).value or "").strip() == tc_col_name:
                return row, col
    raise ValueError(f"'{tc_col_name}' 헤더를 찾지 못했습니다.")


def write_results(source_xlsx, results, out_path=None, sheet_name=None):
    """판정 결과를 체크리스트에 기록하고 저장 경로를 반환한다.

    results: TCResult 리스트. TC ID로 행을 매칭한다.
    체크리스트에 없는 TC ID는 시트 끝에 '자동화 추가 항목'으로 덧붙인다.
    """
    if not os.path.isfile(source_xlsx):
        raise FileNotFoundError(source_xlsx)

    out_path = out_path or os.path.join(
        os.path.dirname(source_xlsx) or ".",
        f"Checklist_Result_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    shutil.copyfile(source_xlsx, out_path)

    wb = load_workbook(out_path)
    ws = wb[sheet_name] if sheet_name else (
        wb[CHECKLIST_SHEET] if CHECKLIST_SHEET in wb.sheetnames
        else _pick_tc_sheet(wb))
    hdr_row, tc_col = _find_header_row(ws)

    # 결과 열 확보 (이미 있으면 재사용)
    existing = {str(ws.cell(hdr_row, c).value or "").strip(): c
                for c in range(1, ws.max_column + 1)}
    first_new = ws.max_column + 1
    col_of = {}
    for i, name in enumerate(RESULT_HEADERS):
        col_of[name] = existing.get(name, first_new + i)
        cell = ws.cell(hdr_row, col_of[name], name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    by_id = {}
    for r in results:
        by_id.setdefault(r.tc_id, []).append(r)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    written, matched_ids = 0, set()

    for row in range(hdr_row + 1, ws.max_row + 1):
        tc_id = str(ws.cell(row, tc_col).value or "").strip()
        if not tc_id:
            continue
        hits = by_id.get(tc_id)
        if not hits:
            _set(ws, row, col_of["자동화 판정"], "미수행")
            continue
        matched_ids.add(tc_id)
        _write_row(ws, row, col_of, hits, stamp)
        written += 1

    # 체크리스트에 없는 TC (예: _mid 보조 판정)
    extra = [tid for tid in by_id if tid not in matched_ids]
    if extra:
        row = ws.max_row + 2
        ws.cell(row, tc_col, "자동화 추가 항목").font = Font(bold=True)
        for tid in sorted(extra):
            row += 1
            ws.cell(row, tc_col, tid)
            _write_row(ws, row, col_of, by_id[tid], stamp)

    for name in RESULT_HEADERS:
        ws.column_dimensions[
            ws.cell(hdr_row, col_of[name]).column_letter].width = (
            12 if name in ("자동화 판정", *STATUSES) else 40)

    wb.save(out_path)
    return {"path": out_path, "written": written, "extra": len(extra),
            "sheet": ws.title}


def _pick_tc_sheet(wb):
    for ws in wb:
        for row in range(1, min(ws.max_row, 20) + 1):
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(row, col).value or "").strip() == "TC ID":
                    return ws
    return wb.active


def _set(ws, row, col, value, status=None):
    cell = ws.cell(row, col, value)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    key = status or (value if value in FILLS else None)
    if key in FILLS:
        cell.fill = FILLS[key]
        cell.font = FONTS[key]
    return cell


def _write_row(ws, row, col_of, hits, stamp):
    verdicts = [h.verdict for h in hits]
    verdict = (FAIL if FAIL in verdicts else
               BLOCKED if BLOCKED in verdicts else
               MANUAL if MANUAL in verdicts else
               PASS if PASS in verdicts else SKIP)

    total = {status: 0 for status in STATUSES}
    fails, manuals, evidence = [], [], []
    for h in hits:
        for k, v in h.counts.items():
            total[k] += v
        for c in h.checks:
            if c.status == FAIL:
                fails.append(f"[Step {c.step}] {c.title} — 기대={c.expected} / 실제={c.actual}")
            elif c.status in (MANUAL, SKIP, BLOCKED):
                # **사유(note)를 함께 적는다** (2026-08-21 사용자 요청 "비고도 잘
                # 기록해 달라"). 제목만 적으면 체크리스트만 받은 사람이 "왜 수동인가 /
                # 왜 대상이 아닌가"를 알 수 없어 리포트를 다시 열어야 했다.
                # SKIP 은 MANUAL 과 뜻이 다르므로 접두사로 구분한다 —
                # MANUAL = 확인해야 하는데 자동으로 못 함,
                # SKIP   = 이 환경에서 확인 대상이 아님.
                reason = " ".join(str(c.note or "").split())
                manuals.append(
                    f"[{c.status}] [Step {c.step}] {c.title}"
                    + (f" — {reason}" if reason else ""))
        evidence.extend(h.evidence)

    _set(ws, row, col_of["자동화 판정"], verdict, verdict)
    _set(ws, row, col_of["판정 일시"], stamp)
    for k in STATUSES:
        _set(ws, row, col_of[k], total[k])

    for name, items in (("실패 항목", fails), ("수동 확인 / 미수행 항목", manuals),
                        ("증거", evidence)):
        cell = ws.cell(row, col_of[name], "\n".join(items) if items else "")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
